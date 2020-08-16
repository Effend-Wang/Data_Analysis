# ----------------------------------------------------------------------------
# Import local python lib
import math
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment

# Import program lib
import file_operation
import cpk_img
import log

# ----------------------------------------------------------------------------
# Parameter definition
upp_limit_name="Upper Limit"
low_limit_name="Lower Limit"
result_file="CPK_Analysis_Result.xlsx"

# ----------------------------------------------------------------------------
# This is a function to choose data range
def data_range():

    # Need user to input an integer number
    print('Please input the number beginning of row:')
    begin_row=int(input())-1
    print('Please input the number beginning of col:')
    begin_col=int(input())-1
    print("Row begins at %d. Col begins at %d" %(begin_row+1,begin_col+1))

    # Record information of data range
    log.write("info","CPK - Chosen Range:\nRow begins at %d\nCol begins at %d" %(begin_row+1,begin_col+1))

    return begin_row, begin_col

# ----------------------------------------------------------------------------
# This is a function to calculate data average & standard deviaion
def data_cpk_cal(para_name,workdata,upp_limit_row,low_limit_row,limit_col_num,begin_col,cpk_cb):

    # Calculating max & min of data
    single_max=max(workdata)
    single_min=min(workdata)

    # Calculating average of data
    ndata=len(workdata)
    single_avg=sum(workdata)/ndata
    
    # Calculating standard deviaion of data
    single_std=0.0
    for i in range(ndata):
    	single_std=single_std+math.pow(workdata[i]-single_avg,2)
    single_std=math.pow(single_std/(ndata-1),0.5)
    
    # Calculating CPL, CPU, CPK
    upp_limit=upp_limit_row[limit_col_num-begin_col]
    low_limit=low_limit_row[limit_col_num-begin_col]
    if(upp_limit=="NA" and low_limit=="NA"):
        log.write("info","CPK - The limit is [NA,NA]")
    elif(upp_limit=="NA"):
        log.write("info","CPK - The limit is [%s,NA]" %low_limit)
    elif(low_limit=="NA"):
        log.write("info","CPK - The limit is [NA,%s]" %upp_limit)
    else:
        log.write("info","CPK - The limit is [%s,%s]" %(low_limit,upp_limit))

    if(single_std==0 or (upp_limit=="NA" and low_limit=="NA")):
    	log.write("warning","CPK - This parametric std is 0 or limit is NA.")
    	single_cpl="NA"
    	single_cpu="NA"
    	single_cpk="NA"
    elif(upp_limit=="NA"):
    	log.write("warning","CPK - This parametric has no upper limit.")
    	single_cpl=(single_avg-low_limit)/3/single_std
    	single_cpu="NA"
    	single_cpk=single_cpl
    elif(low_limit=="NA"):
    	log.write("warning","CPK - This parametric has no lower limit.")
    	single_cpl="NA"
    	single_cpu=(upp_limit-single_avg)/3/single_std
    	single_cpk=single_cpu
    else:
    	single_cpl=(single_avg-low_limit)/3/single_std
    	single_cpu=(upp_limit-single_avg)/3/single_std
    	single_cpk=min(single_cpl,single_cpu)

    # Output calculating results
    log.write("info","CPK - The STD of this parametric is %f" %single_std)
    log.write("info","CPK - The AVG of this parametric is %f" %single_avg)
    if(single_cpk!="NA"):
        log.write("info","CPK - The CPK of this parametric is %f" %single_cpk)
    else:
        log.write("info","CPK - The CPK of this parametric is NA")

    # Check CPK value
    if(single_cpk=="NA"):
        check_cb=(cpk_cb==1)
        log.write("warning","CPK - Parametric has no CPK value.\n=========================")
    elif(single_cpk<1.33):
        check_cb=True
        log.write("warning","CPK - Parametric CPK is not OK.\n=========================")
    else:
        check_cb=(cpk_cb==1)
        log.write("info","CPK - Parametric CPK is OK.\n=========================")
    
    # Draw images of CPK data
    if(check_cb==True):
        cpk_img.cpkimg_draw(para_name,workdata,low_limit,upp_limit,single_avg,single_std)

    return single_avg,single_std,single_max,single_min,single_cpl,single_cpu,single_cpk

# ----------------------------------------------------------------------------
# This is a function to check CPK values
#def check_value(cpk_value):
#    if(cpk_value=="NA"):
#        log.write("error","CPK - Parametric has no CPK value.\n=========================")
#    elif(cpk_value<1.33):
#        log.write("warning","CPK - Parametric CPK is not OK.\n=========================")
#    else:
#        log.write("info","CPK - Parametric CPK is OK.\n=========================")

# ----------------------------------------------------------------------------
# This is a function to check data, delete blank data
def cpk_data_check(workdata):
    check_data=list(workdata)
    while '' in check_data:
        check_data.remove('')
    new_data=tuple(check_data)
    
    return new_data

# ----------------------------------------------------------------------------
# This is a function to output CPK results to excel file
def output_cpk_result(para,avg,std,pos3_delta,neg3_delta,data_min,data_max,cpl,cpu,cpk,low_limit_row,upp_limit_row):

    # Initial file
    outdata=Workbook()
    outsheet=outdata.worksheets[0]
    outsheet.title='CPK_Result'
    
    # Set style of output data
    # Font Type: Times New Roman
    # Size: 12
    # Border: Thin
    # Alignment: Center
    # Pattern Fill: Red background color when CPK<1.33
    font=Font('Times New Roman',size=12)
    fill=PatternFill(fill_type='solid',start_color='FF0000')
    border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    align=Alignment(horizontal='center',vertical='center')

    # Write excel file
    words=('Parameter','Upper Limit','Lower Limit','AVG','STD','+3Del','-3Del','Max','Min','CPL','CPU','CPK')
    npara=len(avg)
    nwords=len(words)
    for i in range(nwords):
        outsheet.cell(row=i+2,column=2).value=words[i]
    for i in range(npara):
        outsheet.cell(row=2,column=i+3).value=para[i]
        outsheet.cell(row=3,column=i+3).value=upp_limit_row[i]
        outsheet.cell(row=4,column=i+3).value=low_limit_row[i]
        outsheet.cell(row=5,column=i+3).value=avg[i]
        outsheet.cell(row=6,column=i+3).value=std[i]
        outsheet.cell(row=7,column=i+3).value=pos3_delta[i]
        outsheet.cell(row=8,column=i+3).value=neg3_delta[i]
        outsheet.cell(row=9,column=i+3).value=data_max[i]
        outsheet.cell(row=10,column=i+3).value=data_min[i]
        outsheet.cell(row=11,column=i+3).value=cpl[i]
        outsheet.cell(row=12,column=i+3).value=cpu[i]
        outsheet.cell(row=13,column=i+3).value=cpk[i]
        if(cpk[i]!="NA" and cpk[i]<1.33):
            outsheet.cell(row=13,column=i+3).fill=fill
    # Apply style for sheet
    for i in range(nwords):
        for j in range(len(cpk)+1):
            outsheet.cell(row=i+2,column=j+2).font=font
            outsheet.cell(row=i+2,column=j+2).border=border
            outsheet.cell(row=i+2,column=j+2).alignment=align

    print("Output cpk excel file.")
    log.write("info","CPK - Output CPK Excel File.")
    outdata.save(result_file)
    file_operation.result_file_move(result_file)

# ----------------------------------------------------------------------------
# This is a function to analyse CPK
def cpk_analysis():
    print("CPK Analysis:")
    avg=()
    std=()
    pos3_delta=()
    neg3_delta=()
    data_min=()
    data_max=()
    cpl=()
    cpu=()
    cpk=()
    (worksheet,nrows,ncols)=file_operation.file_read()
    (begin_row,begin_col)=data_range()

    # Get first col & row data from sheet
    (upp_row_num,upp_col_num)=file_operation.find_value_by_row(worksheet,nrows,upp_limit_name,0,None)
    (low_row_num,low_col_num)=file_operation.find_value_by_row(worksheet,nrows,low_limit_name,0,None)
    upp_limit_row=file_operation.one_row_data(worksheet,upp_row_num,begin_col,None)
    low_limit_row=file_operation.one_row_data(worksheet,low_row_num,begin_col,None)

    # Get parameter data from sheet
    para=worksheet.row_values(0,start_colx=begin_col,end_colx=None)
    log.write("info","CPK - CPK Analysis Details:\n=========================")

    # Get cpk image mode
    cpk_cb=0
    while(cpk_cb!=1 and cpk_cb!=2):
        print("Choose the way to output CPK distribution image:\n\t1. Output All CPK Distribution\n\t2. Output CPK<1.33 Distribution")
        cpk_cb=int(input())
    
    # Calculating by col
    print("Start calculating, please wait...")
    for i in range(ncols): 
        if(i<begin_col):
            continue
        else:
            para_name=file_operation.one_col_data(worksheet,i,0,1)
            workdata=file_operation.one_col_data(worksheet,i,begin_row,None)
            log.write("info","CPK - Analyse parameter: %s" %para_name)
            workdata=cpk_data_check(workdata)
            if(workdata!=''):
                (single_avg,single_std,single_max,single_min,single_cpl,single_cpu,single_cpk)=data_cpk_cal(para_name,workdata,upp_limit_row,low_limit_row,i,begin_col,cpk_cb)
                # Result data are saved as a tuple list in case of no other change
                avg=avg+(single_avg,)
                std=std+(single_std,)
                pos3_delta=pos3_delta+(single_avg+single_std*3,)
                neg3_delta=pos3_delta+(single_avg-single_std*3,)
                data_min=data_min+(single_min,)
                data_max=data_max+(single_max,)
                cpl=cpl+(single_cpl,)
                cpu=cpu+(single_cpu,)
                cpk=cpk+(single_cpk,)
            else:
                log.write("error","CPK - This parameter has no test value!")
                continue

    # Print results to excel file
    print("All data calculate finished!\nOutput results as excel file...")
    output_cpk_result(para,avg,std,pos3_delta,neg3_delta,data_min,data_max,cpl,cpu,cpk,low_limit_row,upp_limit_row)
    print("Excel output finished!\nProgram finished!\n")
    log.write("info","CPK - All CPK analysis steps finished!\nProgram Finished!\n\n")
