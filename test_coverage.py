import file_operation
import logging
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment

# ----------------------------------------------------------------------------
# Parameter definition
upp_limit_name="Higher Limit"
low_limit_name="Lower Limit"
key_name="Key"
result_file="Test_Coverage_Result.xlsx"

# ----------------------------------------------------------------------------
# Set logging config
# Logging level includes: logging.DEBUG, logging.INFO, logging.WARNING, logging.ERROR, logging.CRITICAL
log_path="RunSteps.log"
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(module)s - %(message)s',filename=log_path)

# ----------------------------------------------------------------------------
# This is a function to combine limits to one string
def combine_limit(para,low_data,upp_data):
    limit=()
    for i in range(len(para)):
        if (low_data[i]==""):
            low="NA"
        else:
            low=low_data[i]
        if (upp_data[i]==""):
            upp="NA"
        else:
            upp=upp_data[i]
        new_string=("[%s,%s]" %(low,upp))
        limit=limit+(new_string,)
    return limit

# ----------------------------------------------------------------------------
# This is a function to apply style for sheet
def apply_style(outsheet,row_begin,col_begin,row_end,col_end):

    # Set style of output data
    # Font Type: Times New Roman
    # Size: 12
    # Border: Thin
    # Alignment: Center
    # Pattern Fill: Red background color when CPK<1.33
    font=Font('Times New Roman',size=12)
    border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    align=Alignment(horizontal='center',vertical='center')

        # Apply style for sheet
    for i in range((row_end-row_begin+1)):
        for j in range((col_end-col_begin+1)):
            outsheet.cell(row=row_begin+i,column=col_begin+j).font=font
            outsheet.cell(row=row_begin+i,column=col_begin+j).border=border
            outsheet.cell(row=row_begin+i,column=col_begin+j).alignment=align

# ----------------------------------------------------------------------------
# This is a function to calculate test coverage
def test_cov_cal(old_para,old_limit,new_para,new_limit):
    same_para=()
    same_limit=()
    diff_para=()
    diff_limit_old=()
    diff_limit_new=()
    update_para=()
    update_limit=()
    delete_para=()
    delete_limit=()

    # Check new parameter
    logging.info("Start checking new paramter:")
    for i in range(len(new_para)):
        para_find_status=0
        limit_find_status=0
        for j in range(len(old_para)):
            if (new_para[i]==old_para[j]):
                para_find_status=1
                if (new_limit[i]==old_limit[j]):
                    limit_find_status=1
                    continue
        if (para_find_status==0 and limit_find_status==0):
            logging.info("Parameter %s is a new one" %new_para[i])
            update_para=update_para+(new_para[i],)
            update_limit=update_limit+(new_limit[i],)
        elif (para_find_status==1 and limit_find_status==0):
            logging.info("New limit %s for %s" %(new_limit[i],new_para[i]))
            diff_para=diff_para+(new_para[i],)
            diff_limit_old=diff_limit_old+(old_limit[j],)
            diff_limit_new=diff_limit_new+(new_limit[i],)
        elif (para_find_status==1 and limit_find_status==1):
            logging.info("Same parameter %s & limit %s" %(new_para[i],new_limit[i]))
            same_para=same_para+(new_para[i],)
            same_limit=same_limit+(new_limit[i],)
        else:
            continue
    logging.info("New parameter check finished.\n=========================")
    
    # Check old parameter
    logging.info("Start checking old parameter:")
    for i in range(len(old_para)):
        para_find_status=0
        for j in range(len(new_para)):
            if(new_para[j]==old_para[i]):
                para_find_status=1
                continue
        if (para_find_status==0):
            logging.info("Parameter %s is not exist in new file" %old_para[i])
            delete_para=delete_para+(old_para[i],)
            delete_limit=delete_limit+(old_limit[i],)
    logging.info("Old parameter check finished.\n=========================")
    
    # Output results
    print("All data analysis finished!\nOutput results as excel file...")
    test_cov_output(same_para,same_limit,diff_para,diff_limit_old,diff_limit_new,update_para,update_limit,delete_para,delete_limit)

# ----------------------------------------------------------------------------
# This is a function to output results
def test_cov_output(same_para,same_limit,diff_para,diff_limit_old,diff_limit_new,new_para,new_limit,delete_para,delete_limit):

    # Initial file
    outdata=Workbook()
    outsheet=outdata.worksheets[0]
    outsheet.title='Coverage_Result'

    col_begin=2
    row_begin=2
    if (same_para!=()):
        logging.info("Write same parameter in file")
        outsheet.cell(row=row_begin,column=col_begin).value="Same Parameter"
        outsheet.cell(row=row_begin,column=col_begin+1).value="Limit"
        for i in range(len(same_para)):
            outsheet.cell(row=i+row_begin+1,column=col_begin).value=same_para[i]
            outsheet.cell(row=i+row_begin+1,column=col_begin+1).value=same_limit[i]
        apply_style(outsheet,row_begin,col_begin,len(same_para)+row_begin,col_begin+1)
        col_begin=col_begin+3
    if (diff_para!=()):
        logging.info("Write different parameter in file")
        outsheet.cell(row=row_begin,column=col_begin).value="Update Parameter"
        outsheet.cell(row=row_begin,column=col_begin+1).value="Old Limit"
        outsheet.cell(row=row_begin,column=col_begin+2).value="New Limit"
        for i in range(len(diff_para)):
            outsheet.cell(row=i+row_begin+1,column=col_begin).value=diff_para[i]
            outsheet.cell(row=i+row_begin+1,column=col_begin+1).value=diff_limit_old[i]
            outsheet.cell(row=i+row_begin+1,column=col_begin+2).value=diff_limit_new[i]
        apply_style(outsheet,row_begin,col_begin,len(diff_para)+row_begin,col_begin+2)
        col_begin=col_begin+4
    if (new_para!=()):
        logging.info("Write new parameter in file")
        outsheet.cell(row=row_begin,column=col_begin).value="New Parameter"
        outsheet.cell(row=row_begin,column=col_begin+1).value="Limit"
        for i in range(len(new_para)):
            outsheet.cell(row=i+row_begin+1,column=col_begin).value=new_para[i]
            outsheet.cell(row=i+row_begin+1,column=col_begin+1).value=new_limit[i]
        apply_style(outsheet,row_begin,col_begin,len(new_para)+row_begin,col_begin+1)
        col_begin=col_begin+3
    if (delete_para!=()):
        logging.info("Write deleted parameter in file")
        outsheet.cell(row=row_begin,column=col_begin).value="Deleted Parameter"
        outsheet.cell(row=row_begin,column=col_begin+1).value="Limit"
        for i in range(len(delete_para)):
            outsheet.cell(row=i+row_begin+1,column=col_begin).value=delete_para[i]
            outsheet.cell(row=i+row_begin+1,column=col_begin+1).value=delete_limit[i]
        apply_style(outsheet,row_begin,col_begin,len(delete_para)+row_begin,col_begin+1)
        col_begin=col_begin+3

    # Write file
    outdata.save(result_file)
    file_operation.result_file_move(result_file)

# ----------------------------------------------------------------------------
# This is a function to run test coverage analysis
def test_coverage_analysis():

    # Read old & new test parameter file
    print("Test Coverage Analysis:\n1. Input Old Test Parameter File:")
    logging.info("Read old test file:")
    (oldsheet,old_nrows,old_ncols)=file_operation.file_read()

    print("2. Input New Test Parameter File:")
    logging.info("Read new test file:")
    (newsheet,new_nrows,new_ncols)=file_operation.file_read()

    print("All File Loaded.")
    logging.info("All file loaded")

    # Find parameter begin position in sheet
    (old_para_row,old_para_col)=file_operation.find_value_by_col(oldsheet,old_ncols,key_name,0,None)
    (new_para_row,new_para_col)=file_operation.find_value_by_col(newsheet,new_ncols,key_name,0,None)

    # Find upper limit & lower limit begin position in sheet
    (old_upp_row,old_upp_col)=file_operation.find_value_by_row(oldsheet,old_nrows,upp_limit_name,0,None)
    (old_low_row,old_low_col)=file_operation.find_value_by_row(oldsheet,old_nrows,low_limit_name,0,None)
    (new_upp_row,new_upp_col)=file_operation.find_value_by_row(newsheet,new_nrows,upp_limit_name,0,None)
    (new_low_row,new_low_col)=file_operation.find_value_by_row(newsheet,new_nrows,low_limit_name,0,None)

    # Get parameter & limit data
    old_para_data=file_operation.one_col_data(oldsheet,old_para_col,old_para_row+1,None)
    logging.info("Old file has %s parameters" %(len(old_para_data)))
    old_upp_data=file_operation.one_col_data(oldsheet,old_upp_col,old_upp_row+1,old_upp_row+len(old_para_data)+1)
    old_low_data=file_operation.one_col_data(oldsheet,old_low_col,old_low_row+1,old_low_row+len(old_para_data)+1)
    print("Old file data loaded.")
    logging.info("Old file data loaded")
    new_para_data=file_operation.one_col_data(newsheet,new_para_col,new_para_row+1,None)
    logging.info("Old file has %s parameters" %(len(new_para_data)))
    new_upp_data=file_operation.one_col_data(newsheet,new_upp_col,new_upp_row+1,new_upp_row+len(new_para_data)+1)
    new_low_data=file_operation.one_col_data(newsheet,new_low_col,new_low_row+1,new_low_row+len(new_para_data)+1)
    print("New file data loaded.")
    logging.info("New file data loaded")

    # Combine limits
    old_limit=combine_limit(old_para_data,old_low_data,old_upp_data)
    new_limit=combine_limit(new_para_data,new_low_data,new_upp_data)
    logging.info("Limits combined")

    # Test Coverage Calculating
    print("Start analysis test coverage, please wait...")
    logging.info("Start analysis test coverage\n=========================")
    test_cov_cal(old_para_data,old_limit,new_para_data,new_limit)
    print("Excel output finished!\nProgram finished!\n")
    logging.info("All test coverage analysis steps finished!\nProgram Finished!\n\n")
